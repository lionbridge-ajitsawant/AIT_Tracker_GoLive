"""AIT Tracker Wizard - PM self-service web UI.

Wraps analyze/provision/migrate so a PM can set up a new project tracker
without touching Python or JSON:

  1. Upload the project's Excel tracker (or pick a standard template)
  2. Review what was detected - rename columns, fix types, prune dropdowns
  3. Enter the target SharePoint site
  4. Dry-run preview, then one click to create the lists and migrate the data

Run locally:   python app.py   (or the Start_Tracker_Wizard.bat launcher)
Run centrally: deploy this folder + scripts/ + schema/ to any host
               (e.g. Azure App Service) and PMs just open the URL.
"""

import io
import json
import re
import sys
import threading
import uuid
from contextlib import redirect_stdout
from pathlib import Path

from flask import Flask, redirect, render_template, request, url_for

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "scripts"))

import openpyxl  # noqa: E402

import analyze  # noqa: E402
import migrate  # noqa: E402
import provision  # noqa: E402
from graph_client import find_list, graph_request, make_session  # noqa: E402

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024  # largest tracker seen is ~5 MB
UPLOADS = ROOT / "webapp" / "uploads"
GENERATED = ROOT / "schema" / "generated"
UPLOADS.mkdir(parents=True, exist_ok=True)
GENERATED.mkdir(parents=True, exist_ok=True)

DRAFTS = {}   # draft_id -> {"schema": dict, "source": str}
JOBS = {}     # job_id   -> {"log": [...], "done": bool, "ok": bool}

TYPE_OPTIONS = ["text", "multiline", "choice", "date", "number"]


def facet_to_type(col):
    if "dateTime" in col:
        return "date"
    if "number" in col:
        return "number"
    if "choice" in col:
        return "choice"
    if col.get("text", {}).get("allowMultipleLines"):
        return "multiline"
    return "text"


def type_to_facet(t, choices):
    if t == "date":
        return {"dateTime": {"format": "dateOnly"}}
    if t == "number":
        return {"number": {"decimalPlaces": "automatic"}}
    if t == "choice":
        return {"choice": {"allowTextEntry": True, "choices": choices or ["Option 1"],
                           "displayAs": "dropDownMenu"}}
    if t == "multiline":
        return {"text": {"allowMultipleLines": True}}
    return {"text": {}}


def builtin_templates():
    return sorted(p for p in (ROOT / "schema").glob("*.json"))


def template_catalog():
    """Friendly metadata for each schema, for the picker. Project templates and
    the one-time admin hub are returned separately."""
    projects, hub = [], []
    for p in builtin_templates():
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except (ValueError, OSError):
            continue
        info = data.get("templateInfo", {})
        entry = {
            "stem": p.stem,
            "name": info.get("friendlyName", data.get("project", p.stem)),
            "blurb": info.get("blurb", ""),
            "lists": info.get("lists", ""),
        }
        (hub if info.get("kind") == "hub" else projects).append(entry)
    return projects, hub


@app.route("/")
def index():
    projects, hub = template_catalog()
    return render_template("index.html", projects=projects, hub=hub)


def save_upload(f):
    """Store an uploaded workbook under a sanitized name; None if not an .xlsx."""
    if not f or not f.filename or not f.filename.lower().endswith(".xlsx"):
        return None
    safe = re.sub(r"[^\w. -]", "_", Path(f.filename).name)
    src = UPLOADS / safe
    f.save(src)
    return src


@app.route("/analyze", methods=["POST"])
def do_analyze():
    src = save_upload(request.files.get("workbook"))
    if src is None:
        return redirect(url_for("index"))

    wb_vals = openpyxl.load_workbook(src, read_only=True, data_only=True)
    wb_formulas = openpyxl.load_workbook(src, read_only=True, data_only=False)
    lists, skipped = [], []
    for name in wb_vals.sheetnames:
        if any(h in name.lower() for h in analyze.SKIP_SHEET_HINTS):
            skipped.append(name)
            continue
        spec = analyze.analyze_sheet(wb_vals[name], wb_formulas[name])
        if spec is None:
            skipped.append(name)
        else:
            lists.append(spec)
    wb_vals.close()
    wb_formulas.close()

    draft_id = uuid.uuid4().hex[:10]
    DRAFTS[draft_id] = {
        "schema": {"project": src.stem, "normalize": {}, "lists": lists},
        "source": str(src),
        "skipped": skipped,
    }
    return redirect(url_for("review", draft_id=draft_id))


@app.route("/template", methods=["POST"])
def use_template():
    name = request.form.get("template", "")
    if name not in {p.stem for p in builtin_templates()}:
        return redirect(url_for("index"))
    path = ROOT / "schema" / f"{name}.json"
    schema = json.loads(path.read_text(encoding="utf-8"))
    draft_id = uuid.uuid4().hex[:10]
    src = save_upload(request.files.get("workbook"))
    DRAFTS[draft_id] = {"schema": schema, "source": str(src) if src else None, "skipped": []}
    return redirect(url_for("review", draft_id=draft_id))


@app.route("/review/<draft_id>")
def review(draft_id):
    draft = DRAFTS.get(draft_id)
    if not draft:
        return redirect(url_for("index"))
    view = []
    for li, lst in enumerate(draft["schema"]["lists"]):
        cols = []
        for ci, col in enumerate(lst["columns"]):
            cols.append({
                "i": ci,
                "name": col["name"],
                "header": col.get("_excelHeader", col["name"]),
                "type": facet_to_type(col),
                "choices": "\n".join(col.get("choice", {}).get("choices", [])),
                "filled": col.get("_filled"),
            })
        view.append({
            "i": li,
            "displayName": lst["displayName"],
            "sheet": lst.get("source", {}).get("sheet", ""),
            "rows": lst.get("_dataRows"),
            "derived": [d["column"] for d in lst.get("_skippedDerivedColumns", [])],
            "columns": cols,
        })
    return render_template("review.html", draft_id=draft_id, lists=view,
                           project=draft["schema"].get("project", ""),
                           skipped=draft.get("skipped", []),
                           has_source=bool(draft.get("source")),
                           type_options=TYPE_OPTIONS)


@app.route("/save/<draft_id>", methods=["POST"])
def save(draft_id):
    draft = DRAFTS.get(draft_id)
    if not draft:
        return redirect(url_for("index"))
    form = request.form
    schema = draft["schema"]
    schema["project"] = form.get("project") or schema.get("project", "Project")

    new_lists = []
    for li, lst in enumerate(schema["lists"]):
        if not form.get(f"list-{li}-include"):
            continue
        lst["displayName"] = form.get(f"list-{li}-name") or lst["displayName"]
        kept_cols, kept_map = [], []
        old_map = {name: idx for idx, name in lst.get("source", {}).get("map", [])}
        for ci, col in enumerate(lst["columns"]):
            if not form.get(f"col-{li}-{ci}-include"):
                continue
            old_name = col["name"]
            new_name = re.sub(r"[^A-Za-z0-9]", "", form.get(f"col-{li}-{ci}-name") or old_name) or old_name
            t = form.get(f"col-{li}-{ci}-type", facet_to_type(col))
            choices_raw = form.get(f"col-{li}-{ci}-choices")
            if choices_raw is None:  # field not submitted: keep what was detected
                choices = col.get("choice", {}).get("choices", [])
            else:
                choices = [c.strip() for c in choices_raw.splitlines() if c.strip()]
            if t == "choice" and not choices:
                choices = col.get("choice", {}).get("choices", [])
            new_col = {"name": new_name, "indexed": col.get("indexed", False)}
            new_col.update(type_to_facet(t, choices))
            if not new_col.get("indexed"):
                new_col.pop("indexed")
            kept_cols.append(new_col)
            if old_name in old_map:
                kept_map.append([old_map[old_name], new_name])
        lst["columns"] = kept_cols
        if "source" in lst:
            lst["source"]["map"] = kept_map
        lst.pop("_dataRows", None)
        lst.pop("_skippedDerivedColumns", None)
        new_lists.append(lst)
    schema["lists"] = new_lists

    out = GENERATED / (re.sub(r"[^\w -]", "_", schema["project"]) + ".json")
    out.write_text(json.dumps(schema, indent=2, ensure_ascii=False), encoding="utf-8")
    draft["schema_path"] = str(out)

    # offline dry-run for the preview page
    preview = []
    if draft.get("source"):
        wb = openpyxl.load_workbook(draft["source"], read_only=True, data_only=True)
        normalize = {k.lower(): v for k, v in schema.get("normalize", {}).items()}
        for lst in schema["lists"]:
            if not lst.get("source", {}).get("map") and lst.get("source", {}).get("titleCol") is None:
                preview.append({"name": lst["displayName"], "count": 0, "sample": []})
                continue
            items = migrate.extract_list(wb, lst, normalize)
            preview.append({
                "name": lst["displayName"],
                "count": len(items),
                "sample": [json.dumps(i, ensure_ascii=False)[:300] for i in items[:3]],
            })
        wb.close()
    else:
        preview = [{"name": lst["displayName"], "count": 0, "sample": []} for lst in schema["lists"]]
    draft["preview"] = preview
    return redirect(url_for("launch", draft_id=draft_id))


@app.route("/launch/<draft_id>")
def launch(draft_id):
    draft = DRAFTS.get(draft_id)
    if not draft:
        return redirect(url_for("index"))
    return render_template("launch.html", draft_id=draft_id,
                           project=draft["schema"].get("project", ""),
                           preview=draft.get("preview", []),
                           has_source=bool(draft.get("source")))


def run_job(job, draft, site_url, do_migrate, pm=""):
    log = job["log"]

    class Tee(io.TextIOBase):
        def write(self, s):
            s = s.strip()
            if s:
                log.append(s)
            return len(s)

    try:
        schema = draft["schema"]
        log.append("Signing in to Microsoft Graph ...")
        with redirect_stdout(Tee()):
            session = make_session()

        m = re.match(r"https?://([^/]+)(/.*?)/?$", site_url.strip())
        if not m:
            raise ValueError(f"Could not parse site URL: {site_url}")
        hostname, path = m.group(1), m.group(2)
        from auth import GRAPH
        r = graph_request(session, "GET", f"{GRAPH}/sites/{hostname}:{path}")
        if not r.ok:
            raise RuntimeError(f"Site lookup failed ({r.status_code}): check the URL and permissions")
        site_id = r.json()["id"]
        log.append(f"Site found: {site_url}")

        for spec in schema["lists"]:
            with redirect_stdout(Tee()):
                provision.provision_list(session, site_id, spec)

        # Auto-register this tracker in the AIT Projects hub (unless this IS the
        # hub schema), so cross-project automation picks it up with no per-list setup.
        is_hub = any(s["displayName"] == provision.HUB_LIST for s in schema["lists"])
        if not is_hub:
            from datetime import date
            with redirect_stdout(Tee()):
                provision.register_project(
                    session, site_id,
                    project_name=schema.get("project", schema["lists"][0]["displayName"]),
                    site_url=site_url, pm=pm,
                    service_type=schema.get("serviceType", ""),
                    start_date=date.today().strftime("%Y-%m-%dT00:00:00Z"),
                )

        if do_migrate and draft.get("source"):
            wb = openpyxl.load_workbook(draft["source"], read_only=True, data_only=True)
            normalize = {k.lower(): v for k, v in schema.get("normalize", {}).items()}
            for spec in schema["lists"]:
                items = migrate.extract_list(wb, spec, normalize)
                if not items:
                    log.append(f"{spec['displayName']}: nothing to migrate")
                    continue
                log.append(f"{spec['displayName']}: migrating {len(items)} items ...")
                with redirect_stdout(Tee()):
                    migrate.upload(session, site_id, spec["displayName"], items)
            wb.close()

        log.append("DONE. Open the site and check the new lists: " + site_url)
        job["ok"] = True
    except BaseException as e:  # noqa: BLE001 - report everything to the PM
        log.append(f"ERROR: {e}")
        job["ok"] = False
    finally:
        job["done"] = True


# ----------------------------------------------------------- DEMO MODE
# Simulates the post-creation experience without touching SharePoint:
# real analysis + real migration extraction, rendered in a SharePoint-style
# list UI. For presentations while Graph access awaits IT approval.

GREEN_WORDS = ("completed", "delivered", "done", "approved", "pass", "resolved", "active", "yes")
AMBER_WORDS = ("progress", "pending", "sent", "review", "setup", "onboarding", "mitigating", "follow")
RED_WORDS = ("rework", "blocked", "fail", "rejected", "open", "at risk", "not started", "no")


def pill_class(value):
    v = str(value).lower()
    if any(w in v for w in GREEN_WORDS):
        return "g"
    if any(w in v for w in AMBER_WORDS):
        return "a"
    if any(w in v for w in RED_WORDS):
        return "r"
    return "b"


@app.route("/demo/<draft_id>", methods=["POST"])
def demo_create(draft_id):
    draft = DRAFTS.get(draft_id)
    if not draft:
        return redirect(url_for("index"))
    schema = draft["schema"]
    normalize = {k.lower(): v for k, v in schema.get("normalize", {}).items()}
    demo_lists, log = [], []
    wb = None
    if draft.get("source"):
        wb = openpyxl.load_workbook(draft["source"], read_only=True, data_only=True)
    for spec in schema["lists"]:
        items = []
        if wb is not None and spec.get("source"):
            items = migrate.extract_list(wb, spec, normalize)
        demo_lists.append({"spec": spec, "items": items})
        log.append(f"Creating list ‘{spec['displayName']}’ with "
                   f"{len(spec['columns'])} columns ... OK")
        if items:
            batches = (len(items) + 19) // 20
            log.append(f"Migrating {len(items)} items into ‘{spec['displayName']}’ "
                       f"({batches} batches of 20) ... OK")
    if wb is not None:
        wb.close()
    log.append("Setting default views (All Items, My items, By Status) ... OK")
    log.append("DONE. Tracker is live.")
    draft["demo_lists"] = demo_lists
    draft["demo_log"] = log
    return redirect(url_for("demo_progress", draft_id=draft_id))


@app.route("/demo/<draft_id>/progress")
def demo_progress(draft_id):
    draft = DRAFTS.get(draft_id)
    if not draft or "demo_log" not in draft:
        return redirect(url_for("index"))
    return render_template("demo_progress.html", draft_id=draft_id, log=draft["demo_log"],
                           project=draft["schema"].get("project", "Project"))


@app.route("/demo/<draft_id>/site")
@app.route("/demo/<draft_id>/site/<int:li>")
def demo_site(draft_id, li=0):
    draft = DRAFTS.get(draft_id)
    if not draft or "demo_lists" not in draft:
        return redirect(url_for("index"))
    demo_lists = draft["demo_lists"]
    li = max(0, min(li, len(demo_lists) - 1))
    cur = demo_lists[li]
    spec, items = cur["spec"], cur["items"]
    cols = spec["columns"][:8]

    def cell(fields, col):
        v = fields.get(col["name"])
        if v is None:
            return {"kind": "blank"}
        if "choice" in col:
            return {"kind": "choice", "value": str(v), "cls": pill_class(v),
                    "choices": col["choice"]["choices"]}
        if "dateTime" in col:
            return {"kind": "text", "value": str(v)[:10]}
        return {"kind": "text", "value": str(v)[:48]}

    rows = []
    for f in items[:50]:
        rows.append({
            "title": str(f.get("Title", ""))[:60],
            "cells": [cell(f, c) for c in cols],
            "fields": {k: (str(v)[:200]) for k, v in f.items()},
        })

    def form_field(col):
        if "choice" in col:
            return {"name": col["name"], "kind": "choice", "choices": col["choice"]["choices"]}
        if "dateTime" in col:
            return {"name": col["name"], "kind": "date"}
        if "number" in col:
            return {"name": col["name"], "kind": "number"}
        multiline = col.get("text", {}).get("allowMultipleLines", False)
        return {"name": col["name"], "kind": "multiline" if multiline else "text"}

    return render_template(
        "demo_site.html", draft_id=draft_id, li=li,
        project=draft["schema"].get("project", "Project"),
        tabs=[(i, dl["spec"]["displayName"], len(dl["items"])) for i, dl in enumerate(demo_lists)],
        list_name=spec["displayName"], col_names=[c["name"] for c in cols],
        rows=rows, total=len(items),
        form_fields=[form_field(c) for c in spec["columns"]],
        added=request.args.get("added") == "1",
    )


@app.route("/demo/<draft_id>/site/<int:li>/new", methods=["POST"])
def demo_new_item(draft_id, li):
    draft = DRAFTS.get(draft_id)
    if not draft or "demo_lists" not in draft:
        return redirect(url_for("index"))
    demo_lists = draft["demo_lists"]
    li = max(0, min(li, len(demo_lists) - 1))
    spec = demo_lists[li]["spec"]
    fields = {"Title": (request.form.get("Title") or "").strip() or "(untitled)"}
    for col in spec["columns"]:
        v = (request.form.get(col["name"]) or "").strip()
        if not v:
            continue
        if "number" in col:
            try:
                v = float(v)
            except ValueError:
                continue
        fields[col["name"]] = v
    demo_lists[li]["items"].insert(0, fields)
    return redirect(url_for("demo_site", draft_id=draft_id, li=li, added=1))


@app.route("/run/<draft_id>", methods=["POST"])
def run(draft_id):
    draft = DRAFTS.get(draft_id)
    if not draft:
        return redirect(url_for("index"))
    site_url = request.form.get("site_url", "")
    do_migrate = bool(request.form.get("migrate"))
    pm = request.form.get("pm", "")
    job_id = uuid.uuid4().hex[:10]
    JOBS[job_id] = {"log": [], "done": False, "ok": False}
    t = threading.Thread(target=run_job, args=(JOBS[job_id], draft, site_url, do_migrate, pm),
                         daemon=True)
    t.start()
    return redirect(url_for("progress", job_id=job_id))


@app.route("/progress/<job_id>")
def progress(job_id):
    return render_template("progress.html", job_id=job_id)


@app.route("/progress/<job_id>/log")
def progress_log(job_id):
    job = JOBS.get(job_id, {"log": ["unknown job"], "done": True, "ok": False})
    return {"log": job["log"], "done": job["done"], "ok": job["ok"]}


if __name__ == "__main__":
    import webbrowser
    port = 5050
    threading.Timer(1.0, lambda: webbrowser.open(f"http://127.0.0.1:{port}")).start()
    app.run(host="127.0.0.1", port=port, debug=False)
