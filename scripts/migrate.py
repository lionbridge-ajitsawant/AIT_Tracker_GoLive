"""Generic Excel -> SharePoint Lists migration, driven by a schema file.

Works for ANY project tracker: the schema JSON (see schema/*.json, or generate
a draft for a new workbook with analyze.py) defines the target lists, their
columns, and how source sheet columns map onto them. Conversion rules come
from each column's type facet:

  dateTime -> ISO date          number -> float/int
  choice   -> normalized value  text   -> string (truncated at 255 unless multiline)

Cleaning applied to every cell: whitespace/nbsp stripping, Excel error values
(#N/A, #REF!, ...) dropped, blank rows skipped (stops scanning after 100
consecutive blanks, so million-row SharePoint bloat is never a problem).
Choice values are case-folded to the canonical spelling in the schema, plus
any explicit fixes in the schema's "normalize" map.

Usage:
  python migrate.py --schema ..\\schema\\transcription_tracker.json --source <tracker.xlsx> --dry-run
  python migrate.py --schema ..\\schema\\transcription_tracker.json --source <tracker.xlsx>
  python migrate.py ... --only Tracking         # one list only
"""

import argparse
import json
import sys
from collections import Counter
from datetime import date, datetime, time as dtime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "out"
EXCEL_ERRORS = {"#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!"}

stats = Counter()


# ------------------------------------------------------------- conversions

def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = value.replace("\xa0", " ").strip()
        if not v or v in EXCEL_ERRORS:
            return None
        return v
    return value


def as_text(value, multiline=False):
    v = clean(value)
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        v = v.strftime("%Y-%m-%d")
    v = str(v)
    if not multiline and len(v) > 255:
        v = v[:252] + "..."
        stats["truncated_text"] += 1
    return v


def as_date(value):
    v = clean(value)
    if v is None:
        return None
    if isinstance(v, datetime) or isinstance(v, date):
        return v.strftime("%Y-%m-%dT00:00:00Z")
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v, fmt).strftime("%Y-%m-%dT00:00:00Z")
            except ValueError:
                pass
    stats["unparseable_date"] += 1
    return None


def as_number(value, integer=False):
    v = clean(value)
    if v is None:
        return None
    if isinstance(v, timedelta):
        return round(v.total_seconds(), 2)
    if isinstance(v, dtime):
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        stats["non_numeric"] += 1
        return None
    return int(n) if integer else n


def convert(value, coldef, normalize):
    if "dateTime" in coldef:
        return as_date(value)
    if "number" in coldef:
        return as_number(value, integer=coldef["number"].get("decimalPlaces") == "none")
    if "choice" in coldef:
        v = as_text(value)
        if v is None:
            return None
        fixed = normalize.get(v.lower())
        if fixed is None:
            canon = {c.lower(): c for c in coldef["choice"]["choices"]}
            fixed = canon.get(v.lower(), v)
        if fixed != v:
            stats[f"normalized:{coldef['name']}"] += 1
        return fixed
    multiline = coldef.get("text", {}).get("allowMultipleLines", False)
    return as_text(value, multiline=multiline)


# --------------------------------------------------------------- extraction

def iter_data_rows(ws, first_data_row, stop_after_blank=100):
    blanks = 0
    for row in ws.iter_rows(min_row=first_data_row, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in row):
            blanks = 0
            yield row
        else:
            blanks += 1
            if blanks >= stop_after_blank:
                return


def cell(row, idx):
    return row[idx] if idx < len(row) else None


def extract_list(wb, spec, normalize):
    src = spec["source"]
    if src["sheet"] not in wb.sheetnames:
        print(f"  WARN sheet {src['sheet']!r} not in workbook - skipping {spec['displayName']!r}")
        return []
    ws = wb[src["sheet"]]
    coldefs = {c["name"]: c for c in spec["columns"]}
    require = src.get("requireAny") or []
    items = []
    for row in iter_data_rows(ws, src["firstDataRow"]):
        if require and all(clean(cell(row, i)) is None for i in require):
            continue
        fields = {}
        title_col = src.get("titleCol")
        if title_col is not None:
            fields["Title"] = as_text(cell(row, title_col)) or "(untitled)"
        for col_idx, name in src["map"]:
            v = convert(cell(row, col_idx), coldefs[name], normalize)
            if v is not None:
                fields[name] = v
        items.append(fields)
    return items


# ------------------------------------------------------------------ upload

def upload(session, site_id, list_name, items):
    from auth import GRAPH
    from graph_client import find_list, graph_request

    lst = find_list(session, site_id, list_name)
    if lst is None:
        sys.exit(f"List {list_name!r} not found on the site - run provision.py first.")
    url = f"/sites/{site_id}/lists/{lst['id']}/items"
    done = failed = 0
    for i in range(0, len(items), 20):
        chunk = items[i : i + 20]
        body = {
            "requests": [
                {
                    "id": str(j),
                    "method": "POST",
                    "url": url,
                    "headers": {"Content-Type": "application/json"},
                    "body": {"fields": fields},
                }
                for j, fields in enumerate(chunk)
            ]
        }
        r = graph_request(session, "POST", f"{GRAPH}/$batch", json=body)
        r.raise_for_status()
        for resp in r.json()["responses"]:
            if resp["status"] in (200, 201):
                done += 1
            else:
                failed += 1
                if failed <= 5:
                    print(f"  item failed ({resp['status']}): {json.dumps(resp.get('body'))[:300]}")
        if (i // 20) % 25 == 0:
            print(f"  {list_name}: {done}/{len(items)} uploaded ...")
    print(f"  {list_name}: {done} uploaded, {failed} failed")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--schema", required=True, help="schema JSON (see schema/ or analyze.py output)")
    ap.add_argument("--source", default=None, help="source .xlsx (default: SOURCE_XLSX in .env)")
    ap.add_argument("--only", default=None, help="migrate a single list by displayName")
    ap.add_argument("--dry-run", action="store_true", help="parse + clean only; write out/*.json")
    args = ap.parse_args()

    import os
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
    src = args.source or os.getenv("SOURCE_XLSX")
    if not src:
        sys.exit("No source workbook: pass --source or set SOURCE_XLSX in .env")

    schema = json.loads(Path(args.schema).read_text(encoding="utf-8"))
    normalize = {k.lower(): v for k, v in schema.get("normalize", {}).items()}

    print(f"schema:  {schema.get('project', args.schema)}")
    print(f"source:  {src}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    extracted = {}
    for spec in schema["lists"]:
        if args.only and spec["displayName"].lower() != args.only.lower():
            continue
        items = extract_list(wb, spec, normalize)
        extracted[spec["displayName"]] = items
        print(f"{spec['displayName']}: {len(items)} items extracted")
    wb.close()

    norm = {k: v for k, v in stats.items() if k.startswith("normalized:")}
    if norm:
        print("\nvalue normalizations applied:")
        for k, v in sorted(norm.items()):
            print(f"  {k.split(':', 1)[1]}: {v} value(s)")
    for key in ("truncated_text", "unparseable_date", "non_numeric"):
        if stats[key]:
            print(f"  {key}: {stats[key]} cell(s)")

    if args.dry_run:
        OUT_DIR.mkdir(exist_ok=True)
        for list_name, items in extracted.items():
            out = OUT_DIR / f"{list_name.lower().replace(' ', '_')}_items.json"
            out.write_text(json.dumps(items, indent=1, ensure_ascii=False), encoding="utf-8")
            print(f"wrote {out}")
        print("\ndry run complete - nothing uploaded.")
        return

    from auth import get_site_id
    from graph_client import make_session
    session = make_session()
    site_id = get_site_id(session)
    for list_name, items in extracted.items():
        if items:
            upload(session, site_id, list_name, items)
    print("migration complete.")


if __name__ == "__main__":
    main()
