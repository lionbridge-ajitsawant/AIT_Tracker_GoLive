"""Generate a DRAFT schema JSON for any new project's Excel tracker.

Point this at the workbook a project hands you and it infers, per sheet:
  - the header row (the row that looks most like column titles)
  - a SharePoint column per Excel column, with the type inferred from the data:
      all dates  -> dateTime        all numbers -> number
      few repeating values -> choice (with the observed values as the dropdown)
      long / multi-line text -> multiline text     otherwise -> text
  - formula-derived columns (>=50% formulas) are EXCLUDED from the mapping and
    listed in "_skippedDerivedColumns" - in Lists those become views, calculated
    columns, or Power BI measures instead of stored data.

Sheets that look like dashboards/instructions/dropdown helpers are skipped and
listed in "_skippedSheets".

ALWAYS review the draft before provisioning: rename lists/columns, prune
choice values, fix the title column, delete junk columns.

Usage:
  python analyze.py "C:\\path\\to\\New_Project_Tracker.xlsx" -o ..\\schema\\new_project.json
"""

import argparse
import json
import re
from collections import Counter
from datetime import date, datetime, time as dtime
from pathlib import Path

import openpyxl

SKIP_SHEET_HINTS = ("instruction", "dropdown", "dashboard", "summary", "do not delete", "information")
MAX_SCAN_ROWS = 5000
CHOICE_MAX_DISTINCT = 25
CHOICE_MAX_LEN = 60


def sanitize(name: str, used: set) -> str:
    words = re.findall(r"[A-Za-z0-9]+", str(name))
    base = "".join(w[:1].upper() + w[1:] for w in words)[:30] or "Col"
    if base[0].isdigit():
        base = "C" + base
    out, n = base, 2
    while out.lower() in used:
        out, n = f"{base}{n}", n + 1
    used.add(out.lower())
    return out


def find_header_row(rows) -> int:
    """Pick the row (1-based, among the first 12) with the most non-empty,
    unique, short string cells - that's almost always the header."""
    best_row, best_score = 1, -1
    for i, row in enumerate(rows[:12], start=1):
        strings = [str(v).strip() for v in row if isinstance(v, str) and str(v).strip()]
        formulas = sum(1 for v in row if isinstance(v, str) and v.startswith("="))
        score = len(set(strings)) - formulas * 2 - sum(1 for s in strings if len(s) > 80)
        if score > best_score:
            best_row, best_score = i, score
    return best_row


def infer_column(values):
    """Return a Graph column facet dict for a list of cleaned cell values."""
    if not values:
        return {"text": {}}
    if all(isinstance(v, (datetime, date)) and not isinstance(v, dtime) for v in values):
        return {"dateTime": {"format": "dateOnly"}}
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in values):
        whole = all(float(v).is_integer() for v in values)
        return {"number": {"decimalPlaces": "none" if whole else "automatic"}}
    texts = [str(v).strip() for v in values]
    distinct = Counter(texts)
    if (
        1 <= len(distinct) <= CHOICE_MAX_DISTINCT
        and len(texts) >= 5
        and len(distinct) / len(texts) < 0.5
        and all(len(t) <= CHOICE_MAX_LEN and "\n" not in t for t in distinct)
    ):
        return {
            "choice": {
                "allowTextEntry": True,
                "choices": [v for v, _ in distinct.most_common()],
                "displayAs": "dropDownMenu",
            }
        }
    if any("\n" in t or len(t) > 255 for t in texts):
        return {"text": {"allowMultipleLines": True}}
    return {"text": {}}


def analyze_sheet(ws_vals, ws_formulas):
    rows = []
    for row in ws_vals.iter_rows(min_row=1, max_row=MAX_SCAN_ROWS, values_only=True):
        rows.append(row)
        if len(rows) > 200 and all(
            all(v is None or str(v).strip() == "" for v in r) for r in rows[-100:]
        ):
            rows = rows[:-100]
            break
    if not rows:
        return None
    header_row = find_header_row(rows)
    headers = rows[header_row - 1]
    data = [r for r in rows[header_row:] if any(v is not None and str(v).strip() != "" for v in r)]
    if not any(h is not None and str(h).strip() for h in headers):
        return None

    formula_rows = []
    for row in ws_formulas.iter_rows(
        min_row=header_row + 1, max_row=min(header_row + 200, MAX_SCAN_ROWS), values_only=True
    ):
        formula_rows.append(row)

    used = {"title", "id"}
    columns, mapping, skipped = [], [], []
    title_col = None
    for idx, header in enumerate(headers):
        header = str(header).strip() if header is not None else ""
        vals = [
            r[idx] for r in data
            if idx < len(r) and r[idx] is not None and str(r[idx]).strip() not in ("", "#N/A")
        ]
        if not header:
            if not vals:
                continue
        elif not data:
            # empty template sheet: draft a text column from the header alone
            name = sanitize(header, used)
            if title_col is None:
                title_col = idx
                continue
            columns.append({"name": name, "_excelHeader": header, "_filled": 0, "text": {}})
            mapping.append([idx, name])
            continue
        n_formula = sum(
            1 for r in formula_rows
            if idx < len(r) and isinstance(r[idx], str) and r[idx].startswith("=")
        )
        n_cells = sum(1 for r in formula_rows if idx < len(r) and r[idx] is not None)
        if n_cells and n_formula / n_cells >= 0.5:
            skipped.append({"column": header or f"col{idx}", "reason": "formula-derived",
                            "sample": next((str(r[idx])[:80] for r in formula_rows
                                            if idx < len(r) and isinstance(r[idx], str)
                                            and r[idx].startswith("=")), "")})
            continue
        name = sanitize(header or f"Col{idx + 1}", used)
        facet = infer_column(vals[:2000])
        if title_col is None and "text" in facet and len(set(map(str, vals))) > len(vals) * 0.8:
            title_col = idx
            continue
        col = {"name": name, "_excelHeader": header, "_filled": len(vals)}
        col.update(facet)
        columns.append(col)
        mapping.append([idx, name])

    return {
        "displayName": ws_vals.title,
        "description": f"Imported from sheet {ws_vals.title!r}. DRAFT - review before provisioning.",
        "source": {
            "sheet": ws_vals.title,
            "firstDataRow": header_row + 1,
            "titleCol": title_col,
            "requireAny": [title_col] if title_col is not None else [mapping[0][0]] if mapping else [],
            "map": mapping,
        },
        "columns": columns,
        "_dataRows": len(data),
        "_skippedDerivedColumns": skipped,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", help="path to the new project's tracker .xlsx")
    ap.add_argument("-o", "--output", default=None, help="where to write the draft schema JSON")
    args = ap.parse_args()

    src = Path(args.workbook)
    wb_vals = openpyxl.load_workbook(src, read_only=True, data_only=True)
    wb_formulas = openpyxl.load_workbook(src, read_only=True, data_only=False)

    lists, skipped_sheets = [], []
    for name in wb_vals.sheetnames:
        if any(h in name.lower() for h in SKIP_SHEET_HINTS):
            skipped_sheets.append({"sheet": name, "reason": "looks like a helper/report sheet"})
            continue
        spec = analyze_sheet(wb_vals[name], wb_formulas[name])
        if spec is None:
            skipped_sheets.append({"sheet": name, "reason": "no detectable header/data"})
            continue
        lists.append(spec)
        print(f"sheet {name!r}: {spec['_dataRows']} data rows, "
              f"{len(spec['columns'])} columns mapped, "
              f"{len(spec['_skippedDerivedColumns'])} formula columns skipped "
              f"(header row {spec['source']['firstDataRow'] - 1})")
    wb_vals.close()
    wb_formulas.close()

    draft = {
        "project": src.stem + " (DRAFT - review before use)",
        "normalize": {},
        "lists": lists,
        "_skippedSheets": skipped_sheets,
    }
    out = Path(args.output) if args.output else src.with_suffix(".schema.json")
    out.write_text(json.dumps(draft, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\ndraft schema written to {out}")
    print("review it (names, choices, title column, junk columns), then:")
    print(f"  python provision.py --schema {out}")
    print(f"  python migrate.py --schema {out} --source \"{src}\" --dry-run")


if __name__ == "__main__":
    main()
